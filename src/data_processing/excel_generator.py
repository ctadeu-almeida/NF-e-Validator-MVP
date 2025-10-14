# -*- coding: utf-8 -*-
"""
Excel Generator Module

This module provides comprehensive Excel/spreadsheet generation capabilities
for EDA analysis results, including formatted reports with charts and statistics.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, ScatterChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from typing import Dict, List, Optional, Any
import os
from datetime import datetime

class ExcelGenerator:
    """
    Comprehensive Excel report generator for EDA analysis
    """

    def __init__(self, data: pd.DataFrame, output_dir: str = "reports"):
        """
        Initialize Excel generator

        Args:
            data (pd.DataFrame): The dataset to analyze
            output_dir (str): Directory to save Excel reports
        """
        self.data = data
        self.output_dir = output_dir
        self.numeric_cols = list(data.select_dtypes(include=[np.number]).columns)
        self.categorical_cols = list(data.select_dtypes(include=['object', 'category']).columns)

        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)

    def create_comprehensive_report(self, filename: str = None) -> str:
        """
        Create a comprehensive Excel report with multiple sheets

        Args:
            filename: Custom filename for the report

        Returns:
            str: Path to the generated Excel file
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"EDA_Report_{timestamp}.xlsx"

        filepath = os.path.join(self.output_dir, filename)

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Create multiple sheets
            self._create_overview_sheet(writer)
            self._create_statistical_summary_sheet(writer)
            self._create_correlation_sheet(writer)
            self._create_outlier_analysis_sheet(writer)
            self._create_categorical_analysis_sheet(writer)
            self._create_data_quality_sheet(writer)

            if 'Time' in self.data.columns:
                self._create_temporal_analysis_sheet(writer)

            # Add raw data sheet
            self.data.to_excel(writer, sheet_name='Raw_Data', index=False)

        # Apply formatting after writing
        self._format_workbook(filepath)

        return filepath

    def _create_overview_sheet(self, writer: pd.ExcelWriter):
        """Create overview sheet with dataset summary"""
        overview_data = {
            'Metric': [
                'Dataset Shape (Rows)',
                'Dataset Shape (Columns)',
                'Numeric Columns',
                'Categorical Columns',
                'Total Missing Values',
                'Missing Value Percentage',
                'Memory Usage (MB)',
                'Data Types Count',
                'Duplicate Rows'
            ],
            'Value': [
                f"{self.data.shape[0]:,}",
                f"{self.data.shape[1]}",
                f"{len(self.numeric_cols)}",
                f"{len(self.categorical_cols)}",
                f"{self.data.isnull().sum().sum():,}",
                f"{(self.data.isnull().sum().sum() / (self.data.shape[0] * self.data.shape[1]) * 100):.2f}%",
                f"{self.data.memory_usage(deep=True).sum() / 1024 / 1024:.2f}",
                f"{len(self.data.dtypes.value_counts())}",
                f"{self.data.duplicated().sum():,}"
            ]
        }

        overview_df = pd.DataFrame(overview_data)
        overview_df.to_excel(writer, sheet_name='Overview', index=False)

        # Column information
        col_info = pd.DataFrame({
            'Column_Name': self.data.columns,
            'Data_Type': self.data.dtypes.astype(str),
            'Non_Null_Count': self.data.count(),
            'Null_Count': self.data.isnull().sum(),
            'Null_Percentage': (self.data.isnull().sum() / len(self.data) * 100).round(2),
            'Unique_Values': self.data.nunique()
        })

        col_info.to_excel(writer, sheet_name='Overview', startrow=len(overview_df) + 3, index=False)

    def _create_statistical_summary_sheet(self, writer: pd.ExcelWriter):
        """Create statistical summary sheet for numeric columns"""
        if len(self.numeric_cols) == 0:
            empty_df = pd.DataFrame({'Message': ['No numeric columns found']})
            empty_df.to_excel(writer, sheet_name='Statistical_Summary', index=False)
            return

        # Basic statistics
        stats_df = self.data[self.numeric_cols].describe()
        stats_df.to_excel(writer, sheet_name='Statistical_Summary')

        # Additional statistics
        additional_stats = pd.DataFrame({
            'Column': self.numeric_cols,
            'Skewness': [self.data[col].skew() for col in self.numeric_cols],
            'Kurtosis': [self.data[col].kurtosis() for col in self.numeric_cols],
            'Mode': [self.data[col].mode().iloc[0] if len(self.data[col].mode()) > 0 else np.nan for col in self.numeric_cols],
            'Range': [self.data[col].max() - self.data[col].min() for col in self.numeric_cols],
            'IQR': [self.data[col].quantile(0.75) - self.data[col].quantile(0.25) for col in self.numeric_cols]
        })

        additional_stats.to_excel(writer, sheet_name='Statistical_Summary',
                                startrow=len(stats_df) + 3, index=False)

    def _create_correlation_sheet(self, writer: pd.ExcelWriter):
        """Create correlation analysis sheet"""
        if len(self.numeric_cols) < 2:
            empty_df = pd.DataFrame({'Message': ['Insufficient numeric columns for correlation analysis']})
            empty_df.to_excel(writer, sheet_name='Correlation_Analysis', index=False)
            return

        # Correlation matrix
        corr_matrix = self.data[self.numeric_cols].corr()
        corr_matrix.to_excel(writer, sheet_name='Correlation_Analysis')

        # Strong correlations summary
        strong_correlations = []
        for i in range(len(corr_matrix.columns)):
            for j in range(i+1, len(corr_matrix.columns)):
                corr_val = corr_matrix.iloc[i, j]
                if abs(corr_val) > 0.5:
                    strong_correlations.append({
                        'Variable_1': corr_matrix.columns[i],
                        'Variable_2': corr_matrix.columns[j],
                        'Correlation': corr_val,
                        'Strength': 'Strong' if abs(corr_val) > 0.7 else 'Moderate'
                    })

        if strong_correlations:
            strong_corr_df = pd.DataFrame(strong_correlations)
            strong_corr_df = strong_corr_df.sort_values('Correlation', key=abs, ascending=False)
            strong_corr_df.to_excel(writer, sheet_name='Correlation_Analysis',
                                  startrow=len(corr_matrix) + 3, index=False)

    def _create_outlier_analysis_sheet(self, writer: pd.ExcelWriter):
        """Create outlier analysis sheet"""
        if len(self.numeric_cols) == 0:
            empty_df = pd.DataFrame({'Message': ['No numeric columns for outlier analysis']})
            empty_df.to_excel(writer, sheet_name='Outlier_Analysis', index=False)
            return

        outlier_summary = []

        for col in self.numeric_cols:
            Q1 = self.data[col].quantile(0.25)
            Q3 = self.data[col].quantile(0.75)
            IQR = Q3 - Q1
            lower_bound = Q1 - 1.5 * IQR
            upper_bound = Q3 + 1.5 * IQR

            outliers = self.data[(self.data[col] < lower_bound) | (self.data[col] > upper_bound)]
            outlier_count = len(outliers)
            outlier_percentage = (outlier_count / len(self.data)) * 100

            outlier_summary.append({
                'Column': col,
                'Q1': Q1,
                'Q3': Q3,
                'IQR': IQR,
                'Lower_Bound': lower_bound,
                'Upper_Bound': upper_bound,
                'Outlier_Count': outlier_count,
                'Outlier_Percentage': outlier_percentage,
                'Min_Outlier': outliers[col].min() if outlier_count > 0 else np.nan,
                'Max_Outlier': outliers[col].max() if outlier_count > 0 else np.nan
            })

        outlier_df = pd.DataFrame(outlier_summary)
        outlier_df.to_excel(writer, sheet_name='Outlier_Analysis', index=False)

    def _create_categorical_analysis_sheet(self, writer: pd.ExcelWriter):
        """Create categorical analysis sheet"""
        if len(self.categorical_cols) == 0:
            empty_df = pd.DataFrame({'Message': ['No categorical columns found']})
            empty_df.to_excel(writer, sheet_name='Categorical_Analysis', index=False)
            return

        current_row = 0

        for col in self.categorical_cols:
            # Value counts for each categorical column
            value_counts = self.data[col].value_counts().head(20)
            value_counts_df = pd.DataFrame({
                'Category': value_counts.index,
                'Count': value_counts.values,
                'Percentage': (value_counts.values / len(self.data) * 100).round(2)
            })

            # Add column header
            header_df = pd.DataFrame({'Info': [f'Analysis for column: {col}']})
            header_df.to_excel(writer, sheet_name='Categorical_Analysis',
                             startrow=current_row, index=False)

            value_counts_df.to_excel(writer, sheet_name='Categorical_Analysis',
                                   startrow=current_row + 2, index=False)

            current_row += len(value_counts_df) + 5

    def _create_data_quality_sheet(self, writer: pd.ExcelWriter):
        """Create data quality assessment sheet"""
        quality_metrics = []

        for col in self.data.columns:
            null_count = self.data[col].isnull().sum()
            null_percentage = (null_count / len(self.data)) * 100
            unique_count = self.data[col].nunique()

            quality_score = 100 - null_percentage
            if unique_count == 1:
                quality_score -= 20  # Penalize constant columns
            elif unique_count == len(self.data):
                quality_score -= 10  # Penalize completely unique columns

            quality_metrics.append({
                'Column': col,
                'Data_Type': str(self.data[col].dtype),
                'Null_Count': null_count,
                'Null_Percentage': null_percentage,
                'Unique_Count': unique_count,
                'Unique_Percentage': (unique_count / len(self.data)) * 100,
                'Quality_Score': max(0, quality_score)
            })

        quality_df = pd.DataFrame(quality_metrics)
        quality_df = quality_df.sort_values('Quality_Score', ascending=False)
        quality_df.to_excel(writer, sheet_name='Data_Quality', index=False)

    def _create_temporal_analysis_sheet(self, writer: pd.ExcelWriter):
        """Create temporal analysis sheet if Time column exists"""
        time_col = 'Time'

        # Basic time statistics
        time_stats = {
            'Metric': [
                'Min Time',
                'Max Time',
                'Time Range (seconds)',
                'Total Transactions',
                'Average Time Between Transactions',
                'Median Time Between Transactions'
            ],
            'Value': [
                self.data[time_col].min(),
                self.data[time_col].max(),
                self.data[time_col].max() - self.data[time_col].min(),
                len(self.data),
                self.data[time_col].diff().mean(),
                self.data[time_col].diff().median()
            ]
        }

        time_stats_df = pd.DataFrame(time_stats)
        time_stats_df.to_excel(writer, sheet_name='Temporal_Analysis', index=False)

        # Time-based aggregations
        time_bins = pd.cut(self.data[time_col], bins=20)

        # Create aggregation dictionary based on available columns
        agg_dict = {time_col: 'count'}
        if 'Amount' in self.data.columns:
            agg_dict['Amount'] = ['mean', 'sum']

        time_aggregation = self.data.groupby(time_bins).agg(agg_dict).round(2)

        time_aggregation.to_excel(writer, sheet_name='Temporal_Analysis',
                                startrow=len(time_stats_df) + 3)

    def _format_workbook(self, filepath: str):
        """Apply formatting to the Excel workbook"""
        from openpyxl import load_workbook

        wb = load_workbook(filepath)

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Format headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filepath)

    def create_summary_report(self, filename: str = None) -> str:
        """
        Create a concise summary report

        Args:
            filename: Custom filename for the summary report

        Returns:
            str: Path to the generated summary Excel file
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"EDA_Summary_{timestamp}.xlsx"

        filepath = os.path.join(self.output_dir, filename)

        # Create summary statistics
        summary_data = {
            'Dataset_Info': {
                'Total_Rows': f"{self.data.shape[0]:,}",
                'Total_Columns': self.data.shape[1],
                'Numeric_Columns': len(self.numeric_cols),
                'Categorical_Columns': len(self.categorical_cols),
                'Missing_Values': f"{self.data.isnull().sum().sum():,}",
                'Missing_Percentage': f"{(self.data.isnull().sum().sum() / (self.data.shape[0] * self.data.shape[1]) * 100):.2f}%"
            }
        }

        # Key insights
        insights = []

        # Missing values insight
        if self.data.isnull().sum().sum() > 0:
            high_missing_cols = self.data.columns[self.data.isnull().sum() / len(self.data) > 0.1]
            if len(high_missing_cols) > 0:
                insights.append(f"Columns with >10% missing values: {list(high_missing_cols)}")

        # Correlation insights
        if len(self.numeric_cols) >= 2:
            corr_matrix = self.data[self.numeric_cols].corr()
            strong_corr_count = ((corr_matrix.abs() > 0.7) & (corr_matrix.abs() < 1.0)).sum().sum() // 2
            insights.append(f"Strong correlations (>0.7) found: {strong_corr_count}")

        # Outlier insights
        outlier_counts = {}
        for col in self.numeric_cols:
            Q1 = self.data[col].quantile(0.25)
            Q3 = self.data[col].quantile(0.75)
            IQR = Q3 - Q1
            outliers = self.data[(self.data[col] < Q1 - 1.5 * IQR) |
                               (self.data[col] > Q3 + 1.5 * IQR)]
            outlier_counts[col] = len(outliers)

        high_outlier_cols = [col for col, count in outlier_counts.items()
                           if count / len(self.data) > 0.05]
        if high_outlier_cols:
            insights.append(f"Columns with >5% outliers: {high_outlier_cols}")

        # Create summary DataFrame
        summary_df = pd.DataFrame([summary_data['Dataset_Info']]).T
        summary_df.columns = ['Value']
        summary_df.reset_index(inplace=True)
        summary_df.columns = ['Metric', 'Value']

        insights_df = pd.DataFrame({'Key_Insights': insights})

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            insights_df.to_excel(writer, sheet_name='Summary',
                               startrow=len(summary_df) + 3, index=False)

            # Add top correlations if available
            if len(self.numeric_cols) >= 2:
                corr_matrix = self.data[self.numeric_cols].corr()
                strong_correlations = []
                for i in range(len(corr_matrix.columns)):
                    for j in range(i+1, len(corr_matrix.columns)):
                        corr_val = corr_matrix.iloc[i, j]
                        if abs(corr_val) > 0.5:
                            strong_correlations.append({
                                'Variable_1': corr_matrix.columns[i],
                                'Variable_2': corr_matrix.columns[j],
                                'Correlation': round(corr_val, 3)
                            })

                if strong_correlations:
                    corr_df = pd.DataFrame(strong_correlations)
                    corr_df = corr_df.sort_values('Correlation', key=abs, ascending=False)
                    corr_df.to_excel(writer, sheet_name='Summary',
                                   startrow=len(summary_df) + len(insights_df) + 6, index=False)

        self._format_workbook(filepath)
        return filepath

    def export_analysis_results(self, analysis_results: Dict[str, Any], filename: str = None) -> str:
        """
        Export custom analysis results to Excel

        Args:
            analysis_results: Dictionary containing analysis results
            filename: Custom filename for the export

        Returns:
            str: Path to the generated Excel file
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Analysis_Results_{timestamp}.xlsx"

        filepath = os.path.join(self.output_dir, filename)

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            for sheet_name, data in analysis_results.items():
                if isinstance(data, pd.DataFrame):
                    data.to_excel(writer, sheet_name=sheet_name, index=False)
                elif isinstance(data, dict):
                    df = pd.DataFrame([data]).T
                    df.columns = ['Value']
                    df.reset_index(inplace=True)
                    df.columns = ['Metric', 'Value']
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    # Convert other data types to DataFrame
                    df = pd.DataFrame({'Result': [str(data)]})
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

        self._format_workbook(filepath)
        return filepath